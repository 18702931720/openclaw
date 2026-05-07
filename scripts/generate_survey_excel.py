#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ===== 样式定义 =====
header_font = Font(name='微软雅黑', bold=True, size=12, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
section_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
section_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
sub_header_font = Font(name='微软雅黑', bold=True, size=10)
sub_header_fill = PatternFill(start_color='D6DCE5', end_color='D6DCE5', fill_type='solid')
normal_font = Font(name='微软雅黑', size=10)
answer_font = Font(name='微软雅黑', size=10, color='808080')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
wrap_alignment = Alignment(wrap_text=True, vertical='top')
center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def set_header(ws, row, col, value, font=header_font, fill=header_fill):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    cell.fill = fill
    cell.border = thin_border
    cell.alignment = center_alignment
    return cell

def set_section_header(ws, row, col, value, colspan=4):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = section_font
    cell.fill = section_fill
    cell.border = thin_border
    cell.alignment = center_alignment
    if colspan > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+colspan-1)

def set_question(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = normal_font
    cell.border = thin_border
    cell.alignment = wrap_alignment
    return cell

def set_answer_cell(ws, row, col, placeholder=''):
    cell = ws.cell(row=row, column=col, value=placeholder)
    cell.font = answer_font
    cell.border = thin_border
    cell.alignment = wrap_alignment
    return cell

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

# ===== Sheet 1: CEO参谋 =====
ws1 = wb.active
ws1.title = "CEO参谋"

# 设置列宽
set_col_width(ws1, 1, 8)
set_col_width(ws1, 2, 50)
set_col_width(ws1, 3, 40)
set_col_width(ws1, 4, 15)

# 标题行
ws1.merge_cells('A1:D1')
title_cell = ws1.cell(row=1, column=1, value="高管数字参谋体系 — CEO参谋需求调研表")
title_cell.font = Font(name='微软雅黑', bold=True, size=14)
title_cell.alignment = center_alignment

# 说明行
ws1.merge_cells('A2:D2')
desc_cell = ws1.cell(row=2, column=1, value="填写说明：请 CEO 本人填写，填写时间约 15-20 分钟。填写完成后发给 AI 参谋整理。")
desc_cell.font = Font(name='微软雅黑', size=9, italic=True)
desc_cell.alignment = Alignment(wrap_text=True)

row = 4
set_section_header(ws1, row, 1, "一、高频决策场景", 4)
row += 1
set_header(ws1, row, 1, "题号")
set_header(ws1, row, 2, "问题")
set_header(ws1, row, 3, "答案")
set_header(ws1, row, 4, "优先级")
row += 1

questions_1_1 = [
    ("Q1", "您最常被问到哪 3 类问题？", ""),
    ("Q2", "您每周花最多时间在哪个领域的决策上？（市场/销售/人力/财务/运营/其他）", ""),
    ("Q3", "您最希望 AI 能直接回答的是什么问题？（选 3 个）", ""),
]
for q in questions_1_1:
    set_question(ws1, row, 1, q[0])
    set_question(ws1, row, 2, q[1])
    set_answer_cell(ws1, row, 3, "请填写...")
    set_answer_cell(ws1, row, 4, "高/中/低")
    row += 1

row += 1
set_section_header(ws1, row, 1, "二、情报需求", 4)
row += 1
set_header(ws1, row, 1, "题号")
set_header(ws1, row, 2, "问题")
set_header(ws1, row, 3, "答案")
set_header(ws1, row, 4, "优先级")
row += 1

questions_1_2 = [
    ("Q4", "您每天/每周最想看到哪些数据或指标？（至少 3 个）", ""),
    ("Q5", "您关注哪些竞争对手动态？", ""),
    ("Q6", "您希望在什么时间收到主动推送？推送什么内容？", ""),
]
for q in questions_1_2:
    set_question(ws1, row, 1, q[0])
    set_question(ws1, row, 2, q[1])
    set_answer_cell(ws1, row, 3, "请填写...")
    set_answer_cell(ws1, row, 4, "高/中/低")
    row += 1

row += 1
set_section_header(ws1, row, 1, "三、会议场景", 4)
row += 1
set_header(ws1, row, 1, "题号")
set_header(ws1, row, 2, "问题")
set_header(ws1, row, 3, "答案")
set_header(ws1, row, 4, "优先级")
row += 1

questions_1_3 = [
    ("Q7", "开会前，您最希望 AI 帮您准备什么材料？", ""),
    ("Q8", "会议纪要/决议，您希望 AI 如何跟进？（自动追踪/提醒待办/按需查询）", ""),
]
for q in questions_1_3:
    set_question(ws1, row, 1, q[0])
    set_question(ws1, row, 2, q[1])
    set_answer_cell(ws1, row, 3, "请填写...")
    set_answer_cell(ws1, row, 4, "高/中/低")
    row += 1

row += 1
set_section_header(ws1, row, 1, "四、痛点与期望", 4)
row += 1
set_header(ws1, row, 1, "题号")
set_header(ws1, row, 2, "问题")
set_header(ws1, row, 3, "答案")
set_header(ws1, row, 4, "优先级")
row += 1

questions_1_4 = [
    ("Q9", "您目前获取信息最大的痛点是什么？（信息分散/数据不及时/分析深度不够/其他）", ""),
    ("Q10", "您对 CEO 参谋的最大期望是什么？（准确/快速/有洞察/其他）", ""),
]
for q in questions_1_4:
    set_question(ws1, row, 1, q[0])
    set_question(ws1, row, 2, q[1])
    set_answer_cell(ws1, row, 3, "请填写...")
    set_answer_cell(ws1, row, 4, "高/中/低")
    row += 1

# 设置行高
for r in range(1, row):
    ws1.row_dimensions[r].height = 30


# ===== Sheet 2: 市场顾问 =====
ws2 = wb.create_sheet("市场顾问Bot")

set_col_width(ws2, 1, 8)
set_col_width(ws2, 2, 50)
set_col_width(ws2, 3, 40)
set_col_width(ws2, 4, 15)

ws2.merge_cells('A1:D1')
title_cell = ws2.cell(row=1, column=1, value="市场顾问 Bot — 需求调研表")
title_cell.font = Font(name='微软雅黑', bold=True, size=14)
title_cell.alignment = center_alignment

ws2.merge_cells('A2:D2')
desc_cell = ws2.cell(row=2, column=1, value="填写人：市场负责人 ｜ 填写时间约 10 分钟")
desc_cell.font = Font(name='微软雅黑', size=9, italic=True)

row = 4
set_section_header(ws2, row, 1, "一、核心职责", 4)
row += 1
set_header(ws2, row, 1, "题号")
set_header(ws2, row, 2, "问题")
set_header(ws2, row, 3, "答案")
set_header(ws2, row, 4, "优先级")
row += 1

q2_1 = [
    ("Q1", "市场部最常被问到哪 3 类问题？", ""),
    ("Q2", "CEO/高管最常问的市场问题是？", ""),
    ("Q3", "目前回答这些问题需要查阅哪些数据来源？", ""),
]
for q in q2_1:
    set_question(ws2, row, 1, q[0])
    set_question(ws2, row, 2, q[1])
    set_answer_cell(ws2, row, 3, "请填写...")
    set_answer_cell(ws2, row, 4, "高/中/低")
    row += 1

row += 1
set_section_header(ws2, row, 1, "二、知识资产", 4)
row += 1
set_header(ws2, row, 1, "题号")
set_header(ws2, row, 2, "问题")
set_header(ws2, row, 3, "答案")
set_header(ws2, row, 4, "更新频率")
row += 1

q2_2 = [
    ("Q4", "市场部有哪些可复用的文档/报告？（市场报告/竞品分析/品牌手册/媒体监测）", ""),
    ("Q5", "这些文档/报告的更新频率是？", ""),
    ("Q6", "市场相关的核心数据指标有哪些？（至少 3 个）", ""),
]
for q in q2_2:
    set_question(ws2, row, 1, q[0])
    set_question(ws2, row, 2, q[1])
    set_answer_cell(ws2, row, 3, "请填写...")
    set_answer_cell(ws2, row, 4, "实时/周/月/季度")
    row += 1

row += 1
set_section_header(ws2, row, 1, "三、场景优先级", 4)
row += 1
set_header(ws2, row, 1, "题号")
set_header(ws2, row, 2, "场景")
set_header(ws2, row, 3, "说明")
set_header(ws2, row, 4, "优先级")
row += 1

q2_3 = [
    ("Q7", "竞品动态查询", "查询竞品产品/价格/营销动态"),
    ("Q8", "营销活动效果分析", "分析各渠道营销 ROI"),
    ("Q9", "目标用户画像", "用户特征、行为分析"),
    ("Q10", "市场份额估算", "根据数据估算市场份额变化"),
    ("Q11", "品牌声量监测", "品牌曝光度和口碑分析"),
    ("Q12", "营销预算分配建议", "基于效果数据的预算优化建议"),
]
for q in q2_3:
    set_question(ws2, row, 1, q[0])
    set_question(ws2, row, 2, q[1])
    set_answer_cell(ws2, row, 3, q[2])
    set_answer_cell(ws2, row, 4, "高/中/低")
    row += 1


# ===== Sheet 3: 销售顾问 =====
ws3 = wb.create_sheet("销售顾问Bot")

set_col_width(ws3, 1, 8)
set_col_width(ws3, 2, 50)
set_col_width(ws3, 3, 40)
set_col_width(ws3, 4, 15)

ws3.merge_cells('A1:D1')
title_cell = ws3.cell(row=1, column=1, value="销售顾问 Bot — 需求调研表")
title_cell.font = Font(name='微软雅黑', bold=True, size=14)
title_cell.alignment = center_alignment

ws3.merge_cells('A2:D2')
desc_cell = ws3.cell(row=2, column=1, value="填写人：销售负责人 ｜ 填写时间约 10 分钟")
desc_cell.font = Font(name='微软雅黑', size=9, italic=True)

row = 4
set_section_header(ws3, row, 1, "一、核心职责", 4)
row += 1
set_header(ws3, row, 1, "题号")
set_header(ws3, row, 2, "问题")
set_header(ws3, row, 3, "答案")
set_header(ws3, row, 4, "优先级")
row += 1

q3_1 = [
    ("Q1", "销售部最常被问到哪 3 类问题？", ""),
    ("Q2", "CEO/高管最常问的销售问题是？", ""),
    ("Q3", "目前回答这些问题需要查阅哪些数据来源？", ""),
]
for q in q3_1:
    set_question(ws3, row, 1, q[0])
    set_question(ws3, row, 2, q[1])
    set_answer_cell(ws3, row, 3, "请填写...")
    set_answer_cell(ws3, row, 4, "高/中/低")
    row += 1

row += 1
set_section_header(ws3, row, 1, "二、知识资产", 4)
row += 1
set_header(ws3, row, 1, "题号")
set_header(ws3, row, 2, "问题")
set_header(ws3, row, 3, "答案")
set_header(ws3, row, 4, "更新频率")
row += 1

q3_2 = [
    ("Q4", "销售部有哪些可复用的文档/模板？（销售话术/成功案例/产品报价/竞品对比）", ""),
    ("Q5", "销售核心数据指标有哪些？（至少 3 个）", ""),
    ("Q6", "客户画像更新频率？（实时/周/月/季度）", ""),
]
for q in q3_2:
    set_question(ws3, row, 1, q[0])
    set_question(ws3, row, 2, q[1])
    set_answer_cell(ws3, row, 3, "请填写...")
    set_answer_cell(ws3, row, 4, "实时/周/月/季度")
    row += 1

row += 1
set_section_header(ws3, row, 1, "三、场景优先级", 4)
row += 1
set_header(ws3, row, 1, "题号")
set_header(ws3, row, 2, "场景")
set_header(ws3, row, 3, "说明")
set_header(ws3, row, 4, "优先级")
row += 1

q3_3 = [
    ("Q7", "销售业绩实时查询", "当前销售额/完成率/排名"),
    ("Q8", "客户跟进状态追踪", "重点客户跟单进度"),
    ("Q9", "成交预测/漏斗分析", "基于历史数据的成交预估"),
    ("Q10", "销售政策咨询", "折扣/返利/合同条款查询"),
    ("Q11", "竞品报价对比", "竞品价格策略分析"),
    ("Q12", "销售培训/新人上手", "产品知识/话术/FAQ"),
]
for q in q3_3:
    set_question(ws3, row, 1, q[0])
    set_question(ws3, row, 2, q[1])
    set_answer_cell(ws3, row, 3, q[2])
    set_answer_cell(ws3, row, 4, "高/中/低")
    row += 1


# ===== Sheet 4: 人力顾问 =====
ws4 = wb.create_sheet("人力顾问Bot")

set_col_width(ws4, 1, 8)
set_col_width(ws4, 2, 50)
set_col_width(ws4, 3, 40)
set_col_width(ws4, 4, 15)

ws4.merge_cells('A1:D1')
title_cell = ws4.cell(row=1, column=1, value="人力顾问 Bot — 需求调研表")
title_cell.font = Font(name='微软雅黑', bold=True, size=14)
title_cell.alignment = center_alignment

ws4.merge_cells('A2:D2')
desc_cell = ws4.cell(row=2, column=1, value="填写人：人力负责人 ｜ 填写时间约 10 分钟")
desc_cell.font = Font(name='微软雅黑', size=9, italic=True)

row = 4
set_section_header(ws4, row, 1, "一、核心职责", 4)
row += 1
set_header(ws4, row, 1, "题号")
set_header(ws4, row, 2, "问题")
set_header(ws4, row, 3, "答案")
set_header(ws4, row, 4, "优先级")
row += 1

q4_1 = [
    ("Q1", "人力部最常被问到哪 3 类问题？", ""),
    ("Q2", "CEO/高管最常问的人力问题是？", ""),
    ("Q3", "目前回答这些问题需要查阅哪些数据来源？", ""),
]
for q in q4_1:
    set_question(ws4, row, 1, q[0])
    set_question(ws4, row, 2, q[1])
    set_answer_cell(ws4, row, 3, "请填写...")
    set_answer_cell(ws4, row, 4, "高/中/低")
    row += 1

row += 1
set_section_header(ws4, row, 1, "二、知识资产", 4)
row += 1
set_header(ws4, row, 1, "题号")
set_header(ws4, row, 2, "问题")
set_header(ws4, row, 3, "答案")
set_header(ws4, row, 4, "更新频率")
row += 1

q4_2 = [
    ("Q4", "人力部有哪些可复用的文档？（组织架构图/绩效体系/薪酬制度/招聘JD/员工手册）", ""),
    ("Q5", "人力核心数据指标有哪些？（至少 3 个）", ""),
    ("Q6", "哪些信息涉及隐私/保密，需要严格权限？（薪酬/绩效等）", ""),
]
for q in q4_2:
    set_question(ws4, row, 1, q[0])
    set_question(ws4, row, 2, q[1])
    set_answer_cell(ws4, row, 3, "请填写...")
    set_answer_cell(ws4, row, 4, "高/中/低")
    row += 1

row += 1
set_section_header(ws4, row, 1, "三、场景优先级", 4)
row += 1
set_header(ws4, row, 1, "题号")
set_header(ws4, row, 2, "场景")
set_header(ws4, row, 3, "说明")
set_header(ws4, row, 4, "优先级")
row += 1

q4_3 = [
    ("Q7", "组织架构查询", "公司/部门组织架构图"),
    ("Q8", "人员配置/人力成本分析", "各部门的编制和成本"),
    ("Q9", "招聘进度追踪", "在招职位/候选人状态"),
    ("Q10", "绩效制度咨询", "绩效考核规则/流程"),
    ("Q11", "人才盘点/离职风险预警", "核心员工稳定性分析"),
    ("Q12", "HR流程SOP查询", "入职/离职/转正等流程"),
]
for q in q4_3:
    set_question(ws4, row, 1, q[0])
    set_question(ws4, row, 2, q[1])
    set_answer_cell(ws4, row, 3, q[2])
    set_answer_cell(ws4, row, 4, "高/中/低")
    row += 1


# ===== Sheet 5: 财务顾问 =====
ws5 = wb.create_sheet("财务顾问Bot")

set_col_width(ws5, 1, 8)
set_col_width(ws5, 2, 50)
set_col_width(ws5, 3, 40)
set_col_width(ws5, 4, 15)

ws5.merge_cells('A1:D1')
title_cell = ws5.cell(row=1, column=1, value="财务顾问 Bot — 需求调研表")
title_cell.font = Font(name='微软雅黑', bold=True, size=14)
title_cell.alignment = center_alignment

ws5.merge_cells('A2:D2')
desc_cell = ws5.cell(row=2, column=1, value="填写人：财务负责人 ｜ 填写时间约 10 分钟")
desc_cell.font = Font(name='微软雅黑', size=9, italic=True)

row = 4
set_section_header(ws5, row, 1, "一、核心职责", 4)
row += 1
set_header(ws5, row, 1, "题号")
set_header(ws5, row, 2, "问题")
set_header(ws5, row, 3, "答案")
set_header(ws5, row, 4, "优先级")
row += 1

q5_1 = [
    ("Q1", "财务部最常被问到哪 3 类问题？", ""),
    ("Q2", "CEO/高管最常问的财务问题是？", ""),
    ("Q3", "目前回答这些问题需要查阅哪些数据来源？", ""),
]
for q in q5_1:
    set_question(ws5, row, 1, q[0])
    set_question(ws5, row, 2, q[1])
    set_answer_cell(ws5, row, 3, "请填写...")
    set_answer_cell(ws5, row, 4, "高/中/低")
    row += 1

row += 1
set_section_header(ws5, row, 1, "二、知识资产", 4)
row += 1
set_header(ws5, row, 1, "题号")
set_header(ws5, row, 2, "问题")
set_header(ws5, row, 3, "答案")
set_header(ws5, row, 4, "更新频率")
row += 1

q5_2 = [
    ("Q4", "财务部有哪些可复用的报表？（利润表/资产负债表/现金流量表/预算报告/成本分析/审计报告）", ""),
    ("Q5", "财务数据更新频率？（实时/日/周/月/季度）", ""),
    ("Q6", "哪些财务信息需要最高权限保护？", ""),
]
for q in q5_2:
    set_question(ws5, row, 1, q[0])
    set_question(ws5, row, 2, q[1])
    set_answer_cell(ws5, row, 3, "请填写...")
    set_answer_cell(ws5, row, 4, "实时/日/周/月/季度")
    row += 1

row += 1
set_section_header(ws5, row, 1, "三、场景优先级", 4)
row += 1
set_header(ws5, row, 1, "题号")
set_header(ws5, row, 2, "场景")
set_header(ws5, row, 3, "说明")
set_header(ws5, row, 4, "优先级")
row += 1

q5_3 = [
    ("Q7", "实时财务数据查询", "收入/支出/利润实时概况"),
    ("Q8", "预算执行进度", "各部门预算使用情况"),
    ("Q9", "成本结构分析", "成本构成和优化建议"),
    ("Q10", "现金流预测", "未来几个月现金流预估"),
    ("Q11", "同比/环比趋势分析", "财务指标的变化趋势"),
    ("Q12", "风险预警", "异常指标监控和预警"),
]
for q in q5_3:
    set_question(ws5, row, 1, q[0])
    set_question(ws5, row, 2, q[1])
    set_answer_cell(ws5, row, 3, q[2])
    set_answer_cell(ws5, row, 4, "高/中/低")
    row += 1


# ===== Sheet 6: 运营顾问 =====
ws6 = wb.create_sheet("运营顾问Bot")

set_col_width(ws6, 1, 8)
set_col_width(ws6, 2, 50)
set_col_width(ws6, 3, 40)
set_col_width(ws6, 4, 15)

ws6.merge_cells('A1:D1')
title_cell = ws6.cell(row=1, column=1, value="运营顾问 Bot — 需求调研表")
title_cell.font = Font(name='微软雅黑', bold=True, size=14)
title_cell.alignment = center_alignment

ws6.merge_cells('A2:D2')
desc_cell = ws6.cell(row=2, column=1, value="填写人：运营负责人 ｜ 填写时间约 10 分钟")
desc_cell.font = Font(name='微软雅黑', size=9, italic=True)

row = 4
set_section_header(ws6, row, 1, "一、核心职责", 4)
row += 1
set_header(ws6, row, 1, "题号")
set_header(ws6, row, 2, "问题")
set_header(ws6, row, 3, "答案")
set_header(ws6, row, 4, "优先级")
row += 1

q6_1 = [
    ("Q1", "运营部最常被问到哪 3 类问题？", ""),
    ("Q2", "CEO/高管最常问的运营问题是？", ""),
    ("Q3", "目前回答这些问题需要查阅哪些数据来源？", ""),
]
for q in q6_1:
    set_question(ws6, row, 1, q[0])
    set_question(ws6, row, 2, q[1])
    set_answer_cell(ws6, row, 3, "请填写...")
    set_answer_cell(ws6, row, 4, "高/中/低")
    row += 1

row += 1
set_section_header(ws6, row, 1, "二、知识资产", 4)
row += 1
set_header(ws6, row, 1, "题号")
set_header(ws6, row, 2, "问题")
set_header(ws6, row, 3, "答案")
set_header(ws6, row, 4, "更新频率")
row += 1

q6_2 = [
    ("Q4", "运营部有哪些可复用的文档/SOP？（业务流程/供应商管理/客服话术/质量报告）", ""),
    ("Q5", "运营核心数据指标有哪些？（至少 3 个）", ""),
    ("Q6", "运营数据更新频率？（实时/日/周/月）", ""),
]
for q in q6_2:
    set_question(ws6, row, 1, q[0])
    set_question(ws6, row, 2, q[1])
    set_answer_cell(ws6, row, 3, "请填写...")
    set_answer_cell(ws6, row, 4, "实时/日/周/月")
    row += 1

row += 1
set_section_header(ws6, row, 1, "三、场景优先级", 4)
row += 1
set_header(ws6, row, 1, "题号")
set_header(ws6, row, 2, "场景")
set_header(ws6, row, 3, "说明")
set_header(ws6, row, 4, "优先级")
row += 1

q6_3 = [
    ("Q7", "核心运营指标查询", "当日/周/月的核心 KPI"),
    ("Q8", "供应链/库存状态", "库存周转/供应商交货情况"),
    ("Q9", "客服质量/投诉分析", "客服满意度/投诉处理"),
    ("Q10", "流程效率诊断", "各流程的效率问题诊断"),
    ("Q11", "供应商评估", "供应商绩效和风险评估"),
    ("Q12", "运营风险预警", "异常波动监控和预警"),
]
for q in q6_3:
    set_question(ws6, row, 1, q[0])
    set_question(ws6, row, 2, q[1])
    set_answer_cell(ws6, row, 3, q[2])
    set_answer_cell(ws6, row, 4, "高/中/低")
    row += 1


# ===== Sheet 7: 附加题（汇总） =====
ws7 = wb.create_sheet("附加题-CEO+负责人")

set_col_width(ws7, 1, 8)
set_col_width(ws7, 2, 60)
set_col_width(ws7, 3, 30)

ws7.merge_cells('A1:C1')
title_cell = ws7.cell(row=1, column=1, value="附加调研题 — CEO 和各负责人共同填写")
title_cell.font = Font(name='微软雅黑', bold=True, size=14)
title_cell.alignment = center_alignment

row = 3
set_header(ws7, row, 1, "题号")
set_header(ws7, row, 2, "问题")
set_header(ws7, row, 3, "答案")
row += 1

extra_qs = [
    ("附加1", "如果只能选一个痛点让 AI 最先解决，您选哪个？", ""),
    ("附加2", "您能接受的 AI 回答误差范围是多少？（必须准确/大致可用/有参考价值即可）", ""),
    ("附加3", "您最担心 AI 参谋出现什么问题？（回答不准/泄露机密/回答太慢/其他）", ""),
    ("附加4", "您愿意为这套系统投入多少资源？（人力/预算/时间）", ""),
]
for q in extra_qs:
    set_question(ws7, row, 1, q[0])
    set_question(ws7, row, 2, q[1])
    set_answer_cell(ws7, row, 3, "请填写...")
    row += 1


# ===== 保存 =====
output_path = "/home/admin/.openclaw/workspace/高管数字参谋体系_需求调研表.xlsx"
wb.save(output_path)
print(f"Excel 文件已保存到: {output_path}")
